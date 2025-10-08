import json
import boto3
import logging
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger()
logger.setLevel(logging.INFO)

def lambda_handler(event, context):
    try:
        import os
        bucket_name = os.environ.get('S3_BUCKET_NAME')
        if not bucket_name:
            raise ValueError("S3_BUCKET_NAME not set")
        
        logger.info("Fetching Security Hub findings...")
        s3 = boto3.client('s3')
        sh = boto3.client('securityhub')
        
        findings = []
        paginator = sh.get_paginator('get_findings')
        for page in paginator.paginate(Filters={'RecordState': [{'Value': 'ACTIVE', 'Comparison': 'EQUALS'}]}):
            findings.extend(page['Findings'])
        
        logger.info(f"Found {len(findings)} findings")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Security Hub Findings"
        
        headers = ['Title', 'Severity', 'Resource Type', 'Resource ID', 'Status', 'Account', 'Region']
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        
        for f in findings:
            resources = f.get('Resources', [{}])
            ws.append([
                f.get('Title', 'N/A'),
                f.get('Severity', {}).get('Label', 'N/A'),
                resources[0].get('Type', 'N/A'),
                resources[0].get('Id', 'N/A'),
                f.get('Compliance', {}).get('Status', 'N/A'),
                f.get('AwsAccountId', 'N/A'),
                f.get('Region', 'N/A')
            ])
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        key = f"reports/security_hub_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        s3.put_object(Bucket=bucket_name, Key=key, Body=buffer.getvalue())
        
        logger.info(f"Report uploaded to s3://{bucket_name}/{key}")
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Report generated successfully',
                'report': key,
                'findings_count': len(findings)
            })
        }
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error': str(e)})
        }
