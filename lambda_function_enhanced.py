import json
import boto3
import logging
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger()
logger.setLevel(logging.INFO)

def send_sns_notification(topic_arn, bucket_name, report_key, findings_count, high_severity_count):
    """Send SNS notification when report is generated"""
    try:
        sns = boto3.client('sns')
        
        # Generate presigned URL for the report (valid for 7 days)
        s3 = boto3.client('s3')
        url = s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket_name, 'Key': report_key},
            ExpiresIn=604800  # 7 days
        )
        
        message = f"""
SECURITY HUB REPORT GENERATED

Report Details:
- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}
- Total Findings: {findings_count}
- High/Critical Severity: {high_severity_count}

Download Link (valid for 7 days):
{url}

Summary:
- This report contains all active Security Hub findings
- High/Critical findings require immediate attention
- Review the attached Excel file for detailed analysis

Automated GRC System
"""
        
        subject = f"Security Hub Report - {findings_count} Findings ({high_severity_count} High/Critical)"
        
        response = sns.publish(
            TopicArn=topic_arn,
            Message=message,
            Subject=subject
        )
        
        logger.info(f"SNS notification sent: {response['MessageId']}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send SNS notification: {e}")
        return False

def lambda_handler(event, context):
    try:
        import os
        bucket_name = os.environ.get('S3_BUCKET_NAME')
        sns_topic_arn = os.environ.get('SNS_TOPIC_ARN')
        
        if not bucket_name:
            raise ValueError("S3_BUCKET_NAME not set")
        
        logger.info("Fetching Security Hub findings...")
        s3 = boto3.client('s3')
        sh = boto3.client('securityhub')
        
        findings = []
        high_severity_count = 0
        
        paginator = sh.get_paginator('get_findings')
        for page in paginator.paginate(Filters={'RecordState': [{'Value': 'ACTIVE', 'Comparison': 'EQUALS'}]}):
            page_findings = page['Findings']
            findings.extend(page_findings)
            
            # Count high/critical severity findings
            for f in page_findings:
                severity = f.get('Severity', {}).get('Label', 'INFORMATIONAL')
                if severity in ['HIGH', 'CRITICAL']:
                    high_severity_count += 1
        
        logger.info(f"Found {len(findings)} findings ({high_severity_count} high/critical)")
        
        # Create Excel workbook with multiple sheets
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Main findings sheet
        ws = wb.create_sheet("Security Hub Findings")
        
        headers = ['Title', 'Severity', 'Resource Type', 'Resource ID', 'Status', 'Account', 'Region', 'First Observed', 'Last Observed']
        ws.append(headers)
        
        # Style the header row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        
        # Add findings data
        for f in findings:
            resources = f.get('Resources', [{}])
            first_observed = f.get('FirstObservedAt', 'N/A')
            last_observed = f.get('LastObservedAt', 'N/A')
            
            # Format datetime if available
            if first_observed != 'N/A':
                try:
                    first_observed = datetime.fromisoformat(first_observed.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
                except:
                    pass
            if last_observed != 'N/A':
                try:
                    last_observed = datetime.fromisoformat(last_observed.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
                except:
                    pass
            
            ws.append([
                f.get('Title', 'N/A'),
                f.get('Severity', {}).get('Label', 'N/A'),
                resources[0].get('Type', 'N/A') if resources else 'N/A',
                resources[0].get('Id', 'N/A') if resources else 'N/A',
                f.get('Compliance', {}).get('Status', 'N/A'),
                f.get('AwsAccountId', 'N/A'),
                f.get('Region', 'N/A'),
                first_observed,
                last_observed
            ])
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # Create executive summary sheet
        ws_summary = wb.create_sheet("Executive Summary")
        
        # Summary statistics
        severity_counts = {}
        for f in findings:
            severity = f.get('Severity', {}).get('Label', 'INFORMATIONAL')
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        ws_summary.append(['Metric', 'Count'])
        ws_summary.append(['Total Findings', len(findings)])
        ws_summary.append(['High/Critical Severity', high_severity_count])
        ws_summary.append([])
        ws_summary.append(['Severity Breakdown', ''])
        
        for severity, count in sorted(severity_counts.items()):
            ws_summary.append([severity, count])
        
        # Style summary sheet
        for col in range(1, 3):
            cell = ws_summary.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Upload to S3
        key = f"reports/security_hub_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        s3.put_object(Bucket=bucket_name, Key=key, Body=buffer.getvalue())
        
        logger.info(f"Report uploaded to s3://{bucket_name}/{key}")
        
        # Send SNS notification if topic ARN is configured
        if sns_topic_arn:
            send_sns_notification(sns_topic_arn, bucket_name, key, len(findings), high_severity_count)
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Enhanced report generated successfully',
                'report': key,
                'findings_count': len(findings),
                'high_severity_count': high_severity_count,
                'notification_sent': bool(sns_topic_arn)
            })
        }
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error': str(e)})
        }